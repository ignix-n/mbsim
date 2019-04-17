# ===============================================================================
# Title : New Modula Consturction Process Experiment Tool(Simulator)
# Date : 2019.3.12
# Author : Adriano Nam and Gabriel Nam
# Email : adriano97@gmail.com, gabriel9980@naver.com
# Description
#   - Parsing Excel File and Extract Basic Information
#   - Generate a random grouping permutation (integer partition, combination)
#   - Calcuate the group cost and property
#   - Write each group data to a new excel file
# Environment
#   - python3.x
#   - pandas, openpyxl(pip install pandas, pip install openpyxl)
# All copyright reserved to Nam's brothers(Sungyup Nam and Sunghoon Nam)
# ===============================================================================

import sys, os
import re
import openpyxl

from itertools import combinations
from itertools import product
from itertools import chain

# User define library
from filters import GrpShortWall as short_wall
from filters import GrpLongWall as long_wall
from filters import GrpExtWall as ext_wall
from filters import GrpBottom as bottom
from filters import GrpCeiling as ceiling
from filters import GrpPipeRack as pipe_rack
from filters import GrpToiletWall as toilet_wall


class Excel:
  def __init__(self, fname=None):
    self.ws_names = []  # sheet's names (list)

    if fname is not None:
      self.wb = self.open(fname)
      self.ws_names = self.wb.sheetnames
    else:
      self.wb = openpyxl.Workbook()

  def __del__(self):
    self.wb = None
    self.ws_names = []

  def col(self, ws_name, col_num):
    pass

  def open(self, fname):
    return openpyxl.load_workbook(fname)

  def read(self, ws_name):
    """
    :param ws_name: worksheet's name to be read
    :return: list of list for all excel sheets
    ex) [[1, 2, 3, 4], [5, 6, 7, 8]]
    [1, 2, 3, 4] is a row data of excel sheet
    """
    ws = self.wb[ws_name]
    data = []
    for row in ws.values:
      row_data = []
      for value in row:
        row_data.append(value)
      data.append(row_data)
    return data

  def save(self, fname):
    self.wb.save(fname)

  def sheets(self):
    return self.wb.sheetnames

  def write(self, ws_name, data):
    """
    :param ws_name: worksheet name
    :param data: list of list, list type elememt means row data
    :return: none
    """
    if len(self.wb.sheetnames) == 1:
      ws = self.wb.active
      ws.title = ws_name
    else:
      ws = self.wb.create_sheet(ws_name)
    for row_data in data:
      ws.append(row_data)


class IntParti:
  def __init__(self, num):
    self.num = num

  def partition(self, num):
    rst = set()
    rst.add((num,))
    for x in range(1, num):
      for y in self.partition(num - x):
        item = tuple(sorted((x,) + y))
        # if len(item) != self.num:
        #    rst.add(item)
        rst.add(item)
        # all permutation ex (1, 3), (3, 1)
        # rst.add((x, ) + y)
    return rst

  def run(self):
    return self.partition(self.num)


class GrpCombi:
  def __init__(self, elem, grp_seqs, filter):
    """
    :param elem: grouping elements list
    :param grp_size: grouping size list
    """
    self.grp_rst = {}
    self.grp_seqs = grp_seqs
    self.grp_elem = elem
    self.filter = filter

    # init self.combi
    for seq in grp_seqs:
      key_name = '_'.join("%d" % n for n in seq)
      self.grp_rst[key_name] = set()

  def grp2key(self, grp_seq):
    return '_'.join("%d" % n for n in grp_seq)

  def grp_combi(self, grp_rst, grp_left, grp_seq, grp_key):
    # print(grp_rst)
    # print(grp_left)
    # print(grp_seq)
    # print(grp_key)
    if len(grp_seq) == 0:
      self.grp_rst[grp_key].add(tuple(sorted(grp_rst)))
      # print(self.grp_rst[grp_key])
      return

    n = grp_seq.pop(0)

    comb_rst = combinations(grp_left, n)

    for item in comb_rst:
      # print(item)
      if self.filter(item).chk():
        next_rst = list(grp_rst)  # new list
        next_left = set(grp_left)  # new set
        next_seq = list(grp_seq)  # new list

        item = sorted(list(item))
        next_rst.append(tuple(item))
        next_left = next_left - set(item)

        self.grp_combi(next_rst, next_left, next_seq, grp_key)

  def run(self):
    grp_elem = set(self.grp_elem)
    for seq in self.grp_seqs:
      grp_rst = []
      grp_key = self.grp2key(seq)
      self.grp_combi(grp_rst, grp_elem, list(seq), grp_key)
    return self.grp_rst


def sorted_alphanum(l):
  """ Sorts the given iterable in the way that is expected.
  Required arguments:
  l -- The iterable to be sorted.
  """
  convert = lambda text: int(text) if text.isdigit() else text
  alphanum_key = lambda key: [convert(c) for c in re.split('([0-9]+)', key)]
  return sorted(l, key=alphanum_key)


class Simul:
  def __init__(self, fname):
    # set excel file to be analyzed
    dir_path = os.path.dirname(os.path.realpath(__file__))
    self.fname = os.path.join(dir_path, 'input', fname)

    self.xls_data = {}
    self.xls_handler = Excel(self.fname)
    self.xls_ws_name = None

    self.mb_title = []
    self.mb_ids = {}  # key : grp_id
    self.mb_grp = {}  # key : grp_id

    self.proc_title = []
    self.proc_new = []
    self.min_tt = 0  # min total time
    self.min_ta = 0  # min total area
    self.max_merged_len = 0 # max merged length

    self.load_data()

  def constraints(self, grp_id):
    filter = {'SW': short_wall, 'LW': long_wall, 'EW': ext_wall, 'BT': bottom,
              'CL': ceiling, 'PS': pipe_rack, 'TW': toilet_wall}
    return filter[grp_id]

  def load_data(self):
    """
    load three worksheets data into the memory
    :return: None
    """
    ws_names = self.xls_handler.sheets()
    self.xls_ws_name = ws_names  # set worksheet name
    for name in ws_names:
      self.xls_data[name] = self.xls_handler.read(name)

    mb_items = self.xls_data[ws_names[0]]
    self.mb_title = mb_items.pop(0)

    proc_items = self.xls_data[ws_names[1]]
    self.proc_title = proc_items.pop(0)
    # print(self.mb_title)
    # print(mb_items)
    # print(proc_items)
    # print(self.proc_title)

    prev_grp = 'ZZ'
    sub_ids = []
    for item in mb_items:
      # cur_grp, tmp = item[0]
      cur_grp = re.split('(\d+)', item[0])[0]
      if prev_grp != cur_grp:
        if not sub_ids:
          # when sub_item is empty
          sub_ids.append(item[0])
        else:
          if len(sub_ids) != 1:
            self.mb_ids[prev_grp] = sub_ids
          sub_ids = []
          sub_ids.append(item[0])
        prev_grp = cur_grp
      else:
        sub_ids.append(item[0])
    self.mb_ids[prev_grp] = sub_ids
    print(self.mb_ids, file=sys.stderr)

  def ready(self):
    for grp_id, grp_elem in sorted(self.mb_ids.items()):
      elem_num = len(grp_elem)
      grp_seqs = IntParti(elem_num).run()
      grp_combs = GrpCombi(grp_elem, grp_seqs, self.constraints(grp_id)).run()
      # self.mb_grp[grp_id] = grp_combs
      grp_list = []
      for item in chain(grp_combs.values()):
        grp_list.extend(item)
      # self.mb_grp[grp_id] = sorted(grp_combs.values())
      self.mb_grp[grp_id] = grp_list

  def run(self):
    # iterate all cases
    cnt = 1

    self.min_ta = 50000  # init min total area
    self.min_tt = 50000  # init min total time
    self.max_merged_len = 0 # init max merged cell length

    num_elem = 0
    for item in product(*self.mb_grp.values()):
      # print(item) #uncomment if you want to check all combination
      self.update_proc(item)
      self.update_min()
      num_elem += 1
    print("total combination number : {}".format(num_elem), file=sys.stderr)
    print("min area diff: {} min time diff : {}".format(self.min_ta, self.min_tt), file=sys.stderr)
    print("max merged length : {}".format(self.max_merged_len), file=sys.stderr)

    for item in product(*self.mb_grp.values()):
      # print(item)
      self.update_proc(item)
      if self.update_chk():
        fname = "%06d.xlsx" % cnt
        self.proc_new.insert(0, self.proc_title)
        self.save_data(fname, self.proc_new)
        cnt += 1

  def update_proc(self, comb_rule):
    # update modular process table
    self.proc_new = list(self.xls_data[self.xls_ws_name[1]])
    # print(comb_rule)

    for grps in comb_rule:
      for item in grps:
        if len(item) == 1:
          continue
        elif len(item) == 2:
          if 'SW1' in item or 'SW3' in item or 'LW1' in item or 'BT1' in item or 'CL1' in item:
            continue
          else:
            self.update_jobs(item)
        elif len(item) == 3:
          if 'EW1' in item:
            continue
          else:
            self.update_jobs(item)
        else:
          self.update_jobs(item)

  def update_jobs(self, new_mb_ids):
    job_id = []
    time = []
    etime = 0
    time_code = []
    area = []
    del_rows =[]
    for row in self.proc_new:
      if not row[1]:
        continue
      mb_id = row[1].split()[0]
      if mb_id in list(new_mb_ids):
        job_id.append(row[0])
        time.append(row[2])
        time_code.append(row[3])
        area.append(row[4])
        del_rows.append(row)

    #delete the merged row
    for row in del_rows:
      self.proc_new.remove(row)

    if 'BT1' in list(new_mb_ids) or 'SW1' in list(new_mb_ids) or 'CL1' in list(new_mb_ids) or 'EW1' in list(new_mb_ids):
      etime = 0
    elif 'PS1' in list(new_mb_ids) and 'PS2' in list(new_mb_ids):
      etime = 0
    elif 'BT6' in list(new_mb_ids) and 'BT7' in list(new_mb_ids) and 'BT8' in list(new_mb_ids):
      etime = 300
    elif 'TW1' in list(new_mb_ids) and len(list(new_mb_ids)) != 1:
      etime = 420
    else:
      etime = time[0]

    mb_ids = ' '.join(sorted_alphanum(list(new_mb_ids)))
    new_area_code = 'NEW_WA' + ''.join(map(str, job_id))
    # code for handling time field
    # code for handling area field
    new_row = [job_id[0], mb_ids, etime, time_code[0], area[0], new_area_code]

    idx = 0
    new_row_job_id = new_row[0]
    for row in self.proc_new:
          if new_row_job_id < row[0]:
                break
          else:
                idx += 1
          
    #self.proc_new.append(new_row)
    self.proc_new.insert(idx, new_row)


    # print(self.proc_new)
    # print('\n\n')

  def update_min(self):
    """
    process worksheed structure
    row[0] : process id
    row[1] : module id
    row[2] : work time
    row[3] : time code
    row[4] : area
    row[5] : area code
    """
    amount_modula = 120
    deadline = 60 * 8 * 60  # (day * hour * minute)
    default_area = 5000
    area_rate = 0.7
    total_time = 0  # minute
    total_area = 0  # total area
    total_merged_len = 0 # total merged len
    time_counter = 0
    productivity = 0
    all_total_time = 0
    total_area_count = 0
     
    total_area_code = set()

    last_row = None # for productivity
    for row in self.proc_new:
      # proc_id = row[0]
      # if proc_id == 1:
      #  first_time = row[2]
      last_row = list(row)

      # total time
      time_code = row[3]
      if time_code == 'CP':
        total_time += row[2]
        time_counter +=1
   
         
      # total area
      total_area_code.add(row[5])

      # merged new job element length
      if 'NEW' in row[5]: #new module id
        total_merged_len += len(row[1].split()) # the number of id code

    # productivity
    # time_counter = time_code.count('NA')
    productivity = last_row[2]

    # print(total_area_code)
    total_area = len(total_area_code) * 100
    total_area_count = len(total_area_code) 

    all_total_time = total_time + total_area_count*15
    
    print("{} {} {}".format(all_total_time, total_area, total_merged_len))

    # print(productivity)

    time_criteria = all_total_time + productivity * amount_modula
    abs_time_criteria = abs(time_criteria - deadline)

    abs_area_criteria = abs(total_area - default_area * area_rate)

    if int(abs_area_criteria) <= self.min_ta:
      if int(abs_time_criteria) <= self.min_tt:
        if total_merged_len >= self.max_merged_len:
          self.min_ta = int(abs_area_criteria)
          self.min_tt = int(abs_time_criteria)
          self.max_merged_len = total_merged_len

  def update_chk(self):
    """
    process worksheed structure
    row[0] : process id
    row[1] : module id
    row[2] : work time
    row[3] : time code
    row[4] : area
    row[5] : area code
    """
    amount_modula = 120
    deadline = 60 * 8 * 60  # (day * hour * minute)
    default_area = 5000
    area_rate = 0.7

    total_time = 0  # minute
    total_area = 0  # total area
    first_time = 0  # proc id(1)'s working time
    total_merged_len = 0
    time_counter = 0
    productivity = 0
    all_total_time = 0
    total_area_count = 0

    total_area_code = set()

    last_row = None # for productivity
    for row in self.proc_new:
      # proc_id = row[0]
      # if proc_id == 1:
      #  first_time = row[2]
      last_row = list(row)

      # total time
      time_code = row[3]
      if time_code == 'CP':
        total_time += row[2]
        time_counter +=1
     
      # total area
      total_area_code.add(row[5])
      
      
      # merged new job element length
      if 'NEW' in row[5]: #new area code
        total_merged_len += len(row[1].split()) # the number of id code

    # productivity
    # time_counter = time_code.count('NA')
    productivity = last_row[2]

    # print(total_area_code)
    total_area = len(total_area_code) * 100
    total_area_count = len(total_area_code)

    all_total_time = total_time + total_area_count*15

    time_criteria = all_total_time+ productivity * amount_modula
    abs_time_criteria = abs(time_criteria - deadline)

    abs_area_criteria = abs(total_area - default_area * area_rate)

    if int(abs_area_criteria) == int(self.min_ta) and int(abs_time_criteria) == int(self.min_tt) and total_merged_len == self.max_merged_len:
      return True
    else:
      return False

  def save_data(self, fname, data):
    # print(data)
    dir_path = os.path.dirname(os.path.realpath(__file__))
    ofxls = os.path.join(dir_path, 'output', fname)
    xh = Excel()
    xh.write("process", data)
    xh.save(ofxls)
    # print(ofxls)

  def show_new_proc(self):
    for data in self.proc_new:
      print(data)

  def show_process(self):
    ws_name = self.xls_ws_name[1]

    for data in self.xls_data[ws_name]:
      print(data)

  def show_grp_comb(self):
    for key in self.mb_grp.keys():
      print("{}".format(key))
      for mb in self.mb_grp[key]:
        print(mb)

def test():
  grp_elems = ['A1', 'A2', 'A3', 'A4', 'A5']
  grp_seqs = IntParti(5).run()
  grp_rst = GrpCombi(grp_elems, grp_seqs).run()

  print(grp_rst)
  for grp_key, grp_item in grp_rst.items():
    print(grp_key)
    print(len(grp_item))

def run():
  mb_simul = Simul("mb_exp_data.xlsx")
  print('loading data', file=sys.stderr)
  mb_simul.ready()
  print('combination', file=sys.stderr)
  mb_simul.run()
  print("all process done", file=sys.stderr)
  # mb_simul.show_process()
  # print(mb_simul.mb_grp)
  # mb_simul.show_grp_comb()


if __name__ == "__main__":
  run()
